import os
import secrets
from datetime import datetime, date, timedelta
from io import BytesIO
from typing import Optional, List, Dict, Any

import bcrypt
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from apscheduler.schedulers.background import BackgroundScheduler
from fastapi import FastAPI, HTTPException, Depends, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel, Field

from database import get_db_connection, init_postgres_tables

app = FastAPI(title="Financeiro Academia")


def is_postgres() -> bool:
    return bool(os.getenv("DATABASE_URL"))


def get_db():
    conn, _ = get_db_connection()
    return conn


def execute_query(conn, query: str, params: Optional[tuple] = None):
    if is_postgres():
        cur = conn.cursor()
        cur.execute(query.replace("?", "%s"), params or ())
        return cur
    return conn.execute(query, params or ())


def fetchall_dict(conn, query: str, params: Optional[tuple] = None) -> List[Dict[str, Any]]:
    cur = execute_query(conn, query, params)
    rows = cur.fetchall()
    if not rows:
        return []
    if is_postgres():
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in rows]
    return [dict(row) for row in rows]


def fetchone_dict(conn, query: str, params: Optional[tuple] = None):
    cur = execute_query(conn, query, params)
    row = cur.fetchone()
    if not row:
        return None
    if is_postgres():
        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))
    return dict(row)


def scalar(conn, query: str, params: Optional[tuple] = None):
    cur = execute_query(conn, query, params)
    row = cur.fetchone()
    return row[0] if row else 0


def hash_senha(senha: str) -> str:
    return bcrypt.hashpw(senha.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def init_db():
    conn = get_db()
    if is_postgres():
        init_postgres_tables(conn)
    else:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                senha_hash TEXT NOT NULL,
                perfil TEXT DEFAULT 'visualizador',
                ativo INTEGER DEFAULT 1,
                criado_em TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS sessoes (
                token TEXT PRIMARY KEY,
                usuario_id INTEGER,
                expira_em TEXT
            );
            CREATE TABLE IF NOT EXISTS contas_pagar (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                descricao TEXT NOT NULL,
                categoria TEXT,
                valor REAL NOT NULL,
                vencimento TEXT,
                status TEXT DEFAULT 'aberto',
                restrita INTEGER DEFAULT 0,
                comprovante TEXT,
                observacao TEXT,
                numero_doc TEXT,
                centro_custo TEXT,
                forma_pagamento TEXT,
                data_pagamento TEXT,
                recorrente INTEGER DEFAULT 0,
                criado_em TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS contas_receber (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                descricao TEXT NOT NULL,
                categoria TEXT,
                valor REAL NOT NULL,
                vencimento TEXT,
                status TEXT DEFAULT 'aberto',
                restrita INTEGER DEFAULT 0,
                comprovante TEXT,
                observacao TEXT,
                numero_doc TEXT,
                centro_custo TEXT,
                forma_pagamento TEXT,
                data_pagamento TEXT,
                recorrente INTEGER DEFAULT 0,
                criado_em TEXT DEFAULT (datetime('now'))
            );
        """)

    admin_email = os.getenv("ADMIN_EMAIL")
    admin_password = os.getenv("ADMIN_PASSWORD")
    if admin_email and admin_password:
        exists = scalar(conn, "SELECT COUNT(*) FROM usuarios WHERE email=?", (admin_email,))
        if not exists:
            execute_query(conn, "INSERT INTO usuarios (nome,email,senha_hash,perfil,ativo) VALUES (?,?,?,?,1)",
                          ("Administrador", admin_email, hash_senha(admin_password), "admin"))
    conn.commit()
    conn.close()


init_db()


class LoginIn(BaseModel):
    email: str
    senha: str


class UsuarioIn(BaseModel):
    nome: str
    email: str
    senha: str
    perfil: str = "visualizador"


class UsuarioUpdate(BaseModel):
    nome: Optional[str] = None
    email: Optional[str] = None
    senha: Optional[str] = None
    perfil: Optional[str] = None
    ativo: Optional[int] = None


class ContaIn(BaseModel):
    descricao: str
    categoria: Optional[str] = None
    valor: float = Field(gt=0)
    vencimento: str
    status: Optional[str] = "aberto"
    restrita: Optional[int] = 0
    comprovante: Optional[str] = None
    observacao: Optional[str] = None
    numero_doc: Optional[str] = None
    centro_custo: Optional[str] = None
    forma_pagamento: Optional[str] = None
    data_pagamento: Optional[str] = None
    recorrente: Optional[int] = 0


class StatusUpdate(BaseModel):
    status: str


def get_usuario(request: Request):
    token = request.cookies.get("token") or request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        raise HTTPException(401, "Não autenticado")
    conn = get_db()
    try:
        usuario = fetchone_dict(conn,
            "SELECT u.* FROM sessoes s JOIN usuarios u ON u.id=s.usuario_id WHERE s.token=? AND s.expira_em>? AND u.ativo=1",
            (token, datetime.now().isoformat()))
    finally:
        conn.close()
    if not usuario:
        raise HTTPException(401, "Sessão inválida")
    return usuario


def requer_admin(usuario=Depends(get_usuario)):
    if usuario.get("perfil") != "admin":
        raise HTTPException(403, "Acesso restrito")
    return usuario


def pode_editar(usuario=Depends(get_usuario)):
    if usuario.get("perfil") == "visualizador":
        raise HTTPException(403, "Sem permissão para editar")
    return usuario


def ve_restritas(usuario) -> bool:
    return usuario.get("perfil") == "admin"


def filtro_restritas(usuario) -> str:
    return " AND restrita=0" if not ve_restritas(usuario) else ""


@app.post("/api/login")
def login(data: LoginIn):
    conn = get_db()
    try:
        user = fetchone_dict(conn, "SELECT * FROM usuarios WHERE email=? AND ativo=1", (data.email,))
        if not user or not bcrypt.checkpw(data.senha.encode("utf-8"), user["senha_hash"].encode("utf-8")):
            raise HTTPException(401, "E-mail ou senha incorretos")
        token = secrets.token_hex(32)
        expira = (datetime.now() + timedelta(hours=12)).isoformat()
        execute_query(conn, "INSERT INTO sessoes (token,usuario_id,expira_em) VALUES (?,?,?)", (token, user["id"], expira))
        conn.commit()
    finally:
        conn.close()
    resp = JSONResponse({"token": token, "nome": user["nome"], "perfil": user["perfil"]})
    resp.set_cookie("token", token, httponly=True, samesite="lax", max_age=43200)
    return resp


@app.post("/api/logout")
def logout(request: Request):
    token = request.cookies.get("token")
    if token:
        conn = get_db()
        execute_query(conn, "DELETE FROM sessoes WHERE token=?", (token,))
        conn.commit()
        conn.close()
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("token")
    return resp


@app.get("/api/me")
def me(usuario=Depends(get_usuario)):
    return {"id": usuario["id"], "nome": usuario["nome"], "email": usuario["email"], "perfil": usuario["perfil"]}


@app.get("/api/usuarios")
def listar_usuarios(admin=Depends(requer_admin)):
    conn = get_db()
    rows = fetchall_dict(conn, "SELECT id,nome,email,perfil,ativo,criado_em FROM usuarios ORDER BY id")
    conn.close()
    return rows


@app.post("/api/usuarios")
def criar_usuario(data: UsuarioIn, admin=Depends(requer_admin)):
    conn = get_db()
    try:
        execute_query(conn, "INSERT INTO usuarios (nome,email,senha_hash,perfil) VALUES (?,?,?,?)",
                      (data.nome, data.email, hash_senha(data.senha), data.perfil))
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(400, f"Erro ao criar usuário: {e}")
    finally:
        conn.close()
    return {"ok": True}


@app.patch("/api/usuarios/{id}")
def atualizar_usuario(id: int, data: UsuarioUpdate, admin=Depends(requer_admin)):
    conn = get_db()
    campos = []
    params = []
    for campo in ["nome", "email", "perfil", "ativo"]:
        valor = getattr(data, campo)
        if valor is not None:
            campos.append(f"{campo}=?")
            params.append(valor)
    if data.senha:
        campos.append("senha_hash=?")
        params.append(hash_senha(data.senha))
    if campos:
        params.append(id)
        execute_query(conn, f"UPDATE usuarios SET {', '.join(campos)} WHERE id=?", tuple(params))
        conn.commit()
    row = fetchone_dict(conn, "SELECT id,nome,email,perfil,ativo FROM usuarios WHERE id=?", (id,))
    conn.close()
    return row or {"ok": False}


@app.delete("/api/usuarios/{id}")
def deletar_usuario(id: int, admin=Depends(requer_admin)):
    conn = get_db()
    execute_query(conn, "UPDATE usuarios SET ativo=0 WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return {"ok": True}


def listar_contas(tabela: str, status: Optional[str], usuario):
    conn = get_db()
    conds, params = [], []
    if not ve_restritas(usuario):
        conds.append("restrita=0")
    if status and status != "todos":
        conds.append("status=?")
        params.append(status)
    q = f"SELECT * FROM {tabela}"
    if conds:
        q += " WHERE " + " AND ".join(conds)
    q += " ORDER BY vencimento ASC, id DESC"
    rows = fetchall_dict(conn, q, tuple(params))
    conn.close()
    return rows


def criar_conta(tabela: str, conta: ContaIn, usuario):
    if conta.restrita and usuario.get("perfil") != "admin":
        raise HTTPException(403, "Só admin pode criar contas restritas")
    conn = get_db()
    cols = "descricao,categoria,valor,vencimento,status,restrita,comprovante,observacao,numero_doc,centro_custo,forma_pagamento,data_pagamento,recorrente"
    vals = (conta.descricao, conta.categoria, conta.valor, conta.vencimento, conta.status, conta.restrita or 0,
            conta.comprovante, conta.observacao, conta.numero_doc, conta.centro_custo, conta.forma_pagamento,
            conta.data_pagamento, conta.recorrente or 0)
    if is_postgres():
        cur = execute_query(conn, f"INSERT INTO {tabela} ({cols}) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?) RETURNING id", vals)
        new_id = cur.fetchone()[0]
    else:
        cur = execute_query(conn, f"INSERT INTO {tabela} ({cols}) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", vals)
        new_id = cur.lastrowid
    conn.commit()
    row = fetchone_dict(conn, f"SELECT * FROM {tabela} WHERE id=?", (new_id,))
    conn.close()
    return row


@app.get("/api/pagar")
def listar_pagar(status: Optional[str] = None, usuario=Depends(get_usuario)):
    return listar_contas("contas_pagar", status, usuario)


@app.post("/api/pagar")
def criar_pagar(conta: ContaIn, usuario=Depends(pode_editar)):
    return criar_conta("contas_pagar", conta, usuario)


@app.patch("/api/pagar/{id}")
def atualizar_pagar(id: int, update: StatusUpdate, usuario=Depends(pode_editar)):
    conn = get_db()
    execute_query(conn, "UPDATE contas_pagar SET status=? WHERE id=?", (update.status, id))
    conn.commit()
    row = fetchone_dict(conn, "SELECT * FROM contas_pagar WHERE id=?", (id,))
    conn.close()
    return row


@app.delete("/api/pagar/{id}")
def deletar_pagar(id: int, admin=Depends(requer_admin)):
    conn = get_db()
    execute_query(conn, "DELETE FROM contas_pagar WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.get("/api/receber")
def listar_receber(status: Optional[str] = None, usuario=Depends(get_usuario)):
    return listar_contas("contas_receber", status, usuario)


@app.post("/api/receber")
def criar_receber(conta: ContaIn, usuario=Depends(pode_editar)):
    return criar_conta("contas_receber", conta, usuario)


@app.patch("/api/receber/{id}")
def atualizar_receber(id: int, update: StatusUpdate, usuario=Depends(pode_editar)):
    conn = get_db()
    execute_query(conn, "UPDATE contas_receber SET status=? WHERE id=?", (update.status, id))
    conn.commit()
    row = fetchone_dict(conn, "SELECT * FROM contas_receber WHERE id=?", (id,))
    conn.close()
    return row


@app.delete("/api/receber/{id}")
def deletar_receber(id: int, admin=Depends(requer_admin)):
    conn = get_db()
    execute_query(conn, "DELETE FROM contas_receber WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.get("/api/resumo")
def resumo(usuario=Depends(get_usuario)):
    conn = get_db()
    hoje = date.today().isoformat()
    em7 = (date.today() + timedelta(days=7)).isoformat()
    fr = filtro_restritas(usuario)
    a_pagar = float(scalar(conn, f"SELECT COALESCE(SUM(valor),0) FROM contas_pagar WHERE status!='pago'{fr}"))
    a_receber = float(scalar(conn, f"SELECT COALESCE(SUM(valor),0) FROM contas_receber WHERE status!='recebido'{fr}"))
    pago_mes = float(scalar(conn, f"SELECT COALESCE(SUM(valor),0) FROM contas_pagar WHERE status='pago'{fr}"))
    recebido_mes = float(scalar(conn, f"SELECT COALESCE(SUM(valor),0) FROM contas_receber WHERE status='recebido'{fr}"))
    venc7_p = scalar(conn, f"SELECT COUNT(*) FROM contas_pagar WHERE status='aberto' AND vencimento BETWEEN ? AND ?{fr}", (hoje, em7))
    venc7_r = scalar(conn, f"SELECT COUNT(*) FROM contas_receber WHERE status='aberto' AND vencimento BETWEEN ? AND ?{fr}", (hoje, em7))
    vencidos_p = scalar(conn, f"SELECT COUNT(*) FROM contas_pagar WHERE status='aberto' AND vencimento<?{fr}", (hoje,))
    prox = fetchall_dict(conn, f"SELECT *,'pagar' as tipo FROM contas_pagar WHERE status IN ('aberto','vencido') AND vencimento<=?{fr} ORDER BY vencimento LIMIT 5", (em7,))
    prox += fetchall_dict(conn, f"SELECT *,'receber' as tipo FROM contas_receber WHERE status IN ('aberto','vencido') AND vencimento<=?{fr} ORDER BY vencimento LIMIT 5", (em7,))
    conn.close()
    return {"a_pagar": a_pagar, "a_receber": a_receber, "saldo_previsto": a_receber - a_pagar,
            "pago_mes": pago_mes, "recebido_mes": recebido_mes, "venc7_pagar": venc7_p,
            "venc7_receber": venc7_r, "vencidos_pagar": vencidos_p,
            "proximos_vencimentos": sorted(prox, key=lambda x: x.get("vencimento") or "")}


@app.get("/api/dashboard")
def dashboard(usuario=Depends(get_usuario)):
    r = resumo(usuario)
    return {"total_receitas": r["recebido_mes"], "total_despesas": r["pago_mes"], "saldo": r["recebido_mes"] - r["pago_mes"]}


def month_bounds(mes: int, ano: int):
    inicio = date(ano, mes, 1)
    fim = date(ano + (mes // 12), (mes % 12) + 1, 1) - timedelta(days=1)
    return inicio.isoformat(), fim.isoformat()


@app.get("/api/dre")
def get_dre(mes: int = None, ano: int = None, usuario=Depends(get_usuario)):
    hoje = date.today()
    mes = mes or hoje.month
    ano = ano or hoje.year
    inicio, fim = month_bounds(mes, ano)
    fr = filtro_restritas(usuario)
    conn = get_db()
    receitas = fetchall_dict(conn, f"SELECT COALESCE(categoria,'Sem categoria') categoria, COALESCE(SUM(valor),0) total FROM contas_receber WHERE vencimento BETWEEN ? AND ? AND status='recebido'{fr} GROUP BY categoria ORDER BY total DESC", (inicio, fim))
    despesas = fetchall_dict(conn, f"SELECT COALESCE(categoria,'Sem categoria') categoria, COALESCE(SUM(valor),0) total FROM contas_pagar WHERE vencimento BETWEEN ? AND ? AND status='pago'{fr} GROUP BY categoria ORDER BY total DESC", (inicio, fim))
    total_receitas = sum(float(r["total"]) for r in receitas)
    total_despesas = sum(float(r["total"]) for r in despesas)
    labels, comp_r, comp_d = [], [], []
    nomes = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
    for i in range(3, -1, -1):
        m = mes - i
        a = ano
        while m <= 0:
            m += 12
            a -= 1
        ini, fi = month_bounds(m, a)
        labels.append(nomes[m-1])
        comp_r.append(float(scalar(conn, f"SELECT COALESCE(SUM(valor),0) FROM contas_receber WHERE vencimento BETWEEN ? AND ? AND status='recebido'{fr}", (ini, fi))))
        comp_d.append(float(scalar(conn, f"SELECT COALESCE(SUM(valor),0) FROM contas_pagar WHERE vencimento BETWEEN ? AND ? AND status='pago'{fr}", (ini, fi))))
    conn.close()
    return {
        "receitas_por_categoria": {r["categoria"]: float(r["total"]) for r in receitas},
        "total_receitas": total_receitas,
        "despesas_por_categoria": {r["categoria"]: float(r["total"]) for r in despesas},
        "total_despesas": total_despesas,
        "resultado": total_receitas - total_despesas,
        "margem": round(((total_receitas - total_despesas) / total_receitas) * 100, 1) if total_receitas else 0,
        "comparativo_meses": labels,
        "comparativo_receitas": comp_r,
        "comparativo_despesas": comp_d,
    }


@app.get("/api/relatorios")
def gerar_relatorio(tipo: str, formato: str, data_inicio: str, data_fim: str,
                    tipo_conta: str = "todos", status: str = "todos", preview: int = 0, usuario=Depends(get_usuario)):
    """
    Relatórios separados por finalidade:
    - financeiro: listagem detalhada de receitas/despesas no período.
    - dre: resumo por categoria com receitas recebidas e despesas pagas.
    - fluxo: entradas, saídas e saldo acumulado em ordem cronológica.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    tipo = (tipo or "financeiro").lower().strip()
    formato = (formato or "pdf").lower().strip()
    tipo_conta = (tipo_conta or "todos").lower().strip()
    status = (status or "todos").lower().strip()

    tipo_original = tipo
    aliases = {
        "financeiro": "financeiro", "geral": "financeiro", "relatorio": "financeiro", "resumo gerencial": "financeiro",
        "dre": "dre", "resultado": "dre", "demonstrativo": "dre",
        "fluxo": "fluxo", "fluxo de caixa": "fluxo", "fluxo_caixa": "fluxo", "fluxo-de-caixa": "fluxo", "caixa": "fluxo",
        "contas a pagar": "financeiro", "contas_a_pagar": "financeiro", "pagar": "financeiro",
        "contas a receber": "financeiro", "contas_a_receber": "financeiro", "receber": "financeiro",
        "inadimplencia": "financeiro", "inadimplência": "financeiro", "inadimplentes": "financeiro",
        "resumo": "financeiro", "resumo gerencial": "financeiro",
    }
    tipo = aliases.get(tipo, tipo)

    titulos_personalizados = {
        "resumo gerencial": "Resumo Gerencial — Body Fitness",
        "contas a pagar": "Relatório de Contas a Pagar — Body Fitness", "contas_a_pagar": "Relatório de Contas a Pagar — Body Fitness", "pagar": "Relatório de Contas a Pagar — Body Fitness",
        "contas a receber": "Relatório de Contas a Receber — Body Fitness", "contas_a_receber": "Relatório de Contas a Receber — Body Fitness", "receber": "Relatório de Contas a Receber — Body Fitness",
        "inadimplencia": "Relatório de Inadimplência — Body Fitness", "inadimplência": "Relatório de Inadimplência — Body Fitness", "inadimplentes": "Relatório de Inadimplência — Body Fitness",
    }
    titulo_customizado = titulos_personalizados.get(tipo_original)

    if tipo_original in {"contas a pagar", "contas_a_pagar", "pagar"}:
        tipo_conta = "pagar"
    elif tipo_original in {"contas a receber", "contas_a_receber", "receber"}:
        tipo_conta = "receber"
    elif tipo_original in {"inadimplencia", "inadimplência", "inadimplentes"}:
        tipo_conta = "receber"
        if status == "todos":
            status = "aberto"

    if tipo not in {"financeiro", "dre", "fluxo"}:
        raise HTTPException(400, "Tipo de relatório inválido. Use: financeiro, resumo gerencial, fluxo de caixa, DRE, contas a pagar, contas a receber ou inadimplência.")
    formato = "excel" if formato == "xlsx" else formato
    if formato not in {"pdf", "excel"}:
        raise HTTPException(400, "Formato inválido. Use: pdf ou excel.")
    if not data_inicio or not data_fim:
        raise HTTPException(400, "Informe data_inicio e data_fim.")
    if data_inicio > data_fim:
        raise HTTPException(400, "A data inicial não pode ser maior que a data final.")

    def brl(valor):
        return f"R$ {float(valor or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def titulo_relatorio():
        if titulo_customizado:
            return titulo_customizado
        return {
            "financeiro": "Relatório Financeiro — Body Fitness",
            "dre": "DRE — Demonstrativo de Resultado — Body Fitness",
            "fluxo": "Fluxo de Caixa — Body Fitness",
        }[tipo]

    def filename(ext):
        nome = tipo_original.replace(" ", "_").replace("ç", "c").replace("ê", "e").replace("í", "i").replace("ã", "a")
        return f"relatorio_{nome}_{data_inicio}_a_{data_fim}.{ext}"

    fr = filtro_restritas(usuario)
    conn = get_db()

    def carregar_lancamentos(apenas_realizados=False):
        dados = []

        def status_sql_params(status_realizado):
            params = [data_inicio, data_fim]
            sql = ""
            if apenas_realizados:
                sql = " AND status=?"
                params.append(status_realizado)
            elif status and status != "todos":
                sql = " AND status=?"
                params.append(status)
            return sql, tuple(params)

        if tipo_conta in {"todos", "receber", "receitas", "receita"}:
            status_sql, params = status_sql_params("recebido")
            receitas = fetchall_dict(
                conn,
                f"SELECT * FROM contas_receber WHERE vencimento BETWEEN ? AND ?{status_sql}{fr} ORDER BY vencimento, id",
                params,
            )
            dados += [{**r, "tipo": "Receita", "entrada": float(r.get("valor") or 0), "saida": 0.0} for r in receitas]

        if tipo_conta in {"todos", "pagar", "despesas", "despesa"}:
            status_sql, params = status_sql_params("pago")
            despesas = fetchall_dict(
                conn,
                f"SELECT * FROM contas_pagar WHERE vencimento BETWEEN ? AND ?{status_sql}{fr} ORDER BY vencimento, id",
                params,
            )
            dados += [{**r, "tipo": "Despesa", "entrada": 0.0, "saida": float(r.get("valor") or 0)} for r in despesas]

        dados.sort(key=lambda x: (x.get("vencimento") or "", x.get("tipo") or "", x.get("id") or 0))
        return dados

    dados_financeiro = carregar_lancamentos(apenas_realizados=False)
    dados_realizados = carregar_lancamentos(apenas_realizados=True)

    def totais(dados):
        total_receitas = sum(float(d.get("entrada") or 0) for d in dados)
        total_despesas = sum(float(d.get("saida") or 0) for d in dados)
        return total_receitas, total_despesas, total_receitas - total_despesas

    def categorias(dados, tipo_lanc):
        agrupado = {}
        for d in dados:
            if d.get("tipo") != tipo_lanc:
                continue
            cat = d.get("categoria") or "Sem categoria"
            valor = float(d.get("entrada") or d.get("saida") or d.get("valor") or 0)
            agrupado[cat] = agrupado.get(cat, 0.0) + valor
        return sorted(agrupado.items(), key=lambda kv: kv[1], reverse=True)

    dre_receitas = categorias(dados_realizados, "Receita")
    dre_despesas = categorias(dados_realizados, "Despesa")
    dre_total_receitas = sum(v for _, v in dre_receitas)
    dre_total_despesas = sum(v for _, v in dre_despesas)
    dre_resultado = dre_total_receitas - dre_total_despesas
    dre_margem = (dre_resultado / dre_total_receitas * 100) if dre_total_receitas else 0

    fluxo = []
    saldo = 0.0
    for d in dados_financeiro:
        entrada = float(d.get("entrada") or 0)
        saida = float(d.get("saida") or 0)
        saldo += entrada - saida
        fluxo.append({**d, "saldo": saldo})

    if formato == "excel":
        wb = Workbook()
        ws = wb.active

        def style_header(sheet):
            for c in sheet[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="C0392B")
                c.alignment = Alignment(horizontal="center")

        def auto_width(sheet):
            for col in sheet.columns:
                sheet.column_dimensions[col[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in col) + 2, 45)

        if tipo == "financeiro":
            ws.title = "Financeiro"
            ws.append(["Tipo", "Descrição", "Categoria", "Valor", "Vencimento", "Status", "Centro de custo", "Forma pgto", "Observação"])
            style_header(ws)
            for d in dados_financeiro:
                ws.append([d.get("tipo"), d.get("descricao"), d.get("categoria"), float(d.get("valor") or 0), d.get("vencimento"), d.get("status"), d.get("centro_custo"), d.get("forma_pagamento"), d.get("observacao")])
            tr, td, res = totais(dados_financeiro)
            ws.append([])
            ws.append(["Total receitas", "", "", tr])
            ws.append(["Total despesas", "", "", td])
            ws.append(["Resultado", "", "", res])
            auto_width(ws)

        elif tipo == "dre":
            ws.title = "DRE"
            ws.append(["Grupo", "Categoria", "Valor", "% sobre receitas"])
            style_header(ws)
            for cat, valor in dre_receitas:
                ws.append(["Receitas", cat, valor, (valor / dre_total_receitas) if dre_total_receitas else 0])
            for cat, valor in dre_despesas:
                ws.append(["Despesas", cat, valor, (valor / dre_total_receitas) if dre_total_receitas else 0])
            ws.append([])
            ws.append(["TOTAL RECEITAS", "", dre_total_receitas, 1 if dre_total_receitas else 0])
            ws.append(["TOTAL DESPESAS", "", dre_total_despesas, (dre_total_despesas / dre_total_receitas) if dre_total_receitas else 0])
            ws.append(["RESULTADO", "", dre_resultado, (dre_resultado / dre_total_receitas) if dre_total_receitas else 0])
            ws.append(["MARGEM", "", f"{dre_margem:.1f}%", ""])
            auto_width(ws)

        else:
            ws.title = "Fluxo de Caixa"
            ws.append(["Data", "Tipo", "Descrição", "Categoria", "Entrada", "Saída", "Saldo acumulado", "Status", "Forma pgto"])
            style_header(ws)
            for d in fluxo:
                ws.append([d.get("vencimento"), d.get("tipo"), d.get("descricao"), d.get("categoria"), d.get("entrada"), d.get("saida"), d.get("saldo"), d.get("status"), d.get("forma_pagamento")])
            auto_width(ws)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename('xlsx')}"},
        )

    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    small  = ParagraphStyle("small",  parent=styles["Normal"], fontSize=8,  leading=10)
    normal = ParagraphStyle("normal", parent=styles["Normal"], fontSize=9,  leading=12)
    bold9  = ParagraphStyle("bold9",  parent=styles["Normal"], fontSize=9,  leading=12, fontName="Helvetica-Bold")

    def _kpi_table(items):
        row_val = [Paragraph(
            "<font name=\"Helvetica-Bold\" size=\"13\" color=\""+cor+"\">"+val+"</font>",
            styles["Normal"]) for _, val, cor in items]
        row_lbl = [Paragraph(
            "<font size=\"8\" color=\"#888888\">"+lbl+"</font>",
            styles["Normal"]) for lbl, _, __ in items]
        t = Table([row_val, row_lbl], colWidths=[170]*len(items))
        t.setStyle(TableStyle([
            ("ALIGN",        (0,0),(-1,-1),"CENTER"),
            ("VALIGN",       (0,0),(-1,-1),"MIDDLE"),
            ("GRID",         (0,0),(-1,-1),0.5,colors.HexColor("#e0e0e0")),
            ("BACKGROUND",   (0,0),(-1,-1),colors.HexColor("#fafafa")),
            ("TOPPADDING",   (0,0),(-1,-1),8),
            ("BOTTOMPADDING",(0,0),(-1,-1),8),
        ]))
        return t

    def _tbl(data, widths, extra=None):
        base = [
            ("BACKGROUND",    (0,0),(-1,0), colors.HexColor("#C0392B")),
            ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
            ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
            ("GRID",          (0,0),(-1,-1),0.25,colors.HexColor("#cccccc")),
            ("FONTSIZE",      (0,0),(-1,-1),8),
            ("ALIGN",         (2,1),(-1,-1),"RIGHT"),
            ("VALIGN",        (0,0),(-1,-1),"TOP"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f9f9f9")]),
        ]
        if extra: base += extra
        t = Table(data, repeatRows=1, colWidths=widths)
        t.setStyle(TableStyle(base))
        return t

    story = [
        Paragraph(titulo_relatorio(), styles["Title"]),
        Paragraph("Periodo: "+data_inicio+" a "+data_fim+"  |  Gerado em: "+datetime.today().strftime("%d/%m/%Y %H:%M"), styles["Normal"]),
        Spacer(1, 10),
    ]

    if tipo_original in {"resumo", "resumo gerencial"}:
        tr_prev, td_prev, res_prev = totais(dados_financeiro)
        tr_real, td_real, res_real = totais(dados_realizados)
        n_aberto = sum(1 for d in dados_financeiro if (d.get("status") or "")=="aberto")
        story.append(Paragraph("Visao Geral do Periodo", bold9))
        story.append(Spacer(1,4))
        story.append(_kpi_table([
            ("Receitas Previstas",  brl(tr_prev),  "#27AE60"),
            ("Despesas Previstas",  brl(td_prev),  "#C0392B"),
            ("Resultado Previsto",  brl(res_prev), "#27AE60" if res_prev>=0 else "#C0392B"),
            ("Lancamentos em Aberto", str(n_aberto), "#F39C12"),
        ]))
        story.append(Spacer(1,6))
        story.append(_kpi_table([
            ("Receitas Realizadas", brl(tr_real),  "#27AE60"),
            ("Despesas Pagas",      brl(td_real),  "#C0392B"),
            ("Resultado Realizado", brl(res_real), "#27AE60" if res_real>=0 else "#C0392B"),
            ("Total Lancamentos",   str(len(dados_financeiro)), "#2C3E50"),
        ]))
        story.append(Spacer(1,12))
        story.append(Paragraph("Top Categorias - Despesas", bold9)); story.append(Spacer(1,4))
        cat_d = categorias(dados_financeiro,"Despesa")[:8]
        if cat_d:
            rows=[["Categoria","Valor","% do Total"]]
            for cat,val in cat_d: rows.append([cat,brl(val),f"{val/td_prev*100:.1f}%" if td_prev else "-"])
            story.append(_tbl(rows,[320,130,90]))
        else: story.append(Paragraph("Nenhuma despesa no periodo.",normal))
        story.append(Spacer(1,10))
        story.append(Paragraph("Top Categorias - Receitas", bold9)); story.append(Spacer(1,4))
        cat_r = categorias(dados_financeiro,"Receita")[:8]
        if cat_r:
            rows=[["Categoria","Valor","% do Total"]]
            for cat,val in cat_r: rows.append([cat,brl(val),f"{val/tr_prev*100:.1f}%" if tr_prev else "-"])
            story.append(_tbl(rows,[320,130,90]))
        else: story.append(Paragraph("Nenhuma receita no periodo.",normal))

    elif tipo == "dre":
        story.append(_kpi_table([
            ("Receitas Realizadas", brl(dre_total_receitas),"#27AE60"),
            ("Despesas Pagas",      brl(dre_total_despesas),"#C0392B"),
            ("Resultado Liquido",   brl(dre_resultado),     "#27AE60" if dre_resultado>=0 else "#C0392B"),
            ("Margem Liquida",      f"{dre_margem:.1f}%",   "#27AE60" if dre_margem>=0 else "#C0392B"),
        ]))
        story.append(Spacer(1,10))
        story.append(Paragraph("RECEITAS - itens marcados como Recebido", bold9)); story.append(Spacer(1,4))
        if dre_receitas:
            rows=[["Categoria","Valor","% Receitas"]]
            for cat,val in dre_receitas: rows.append([cat,brl(val),f"{val/dre_total_receitas*100:.1f}%" if dre_total_receitas else "0%"])
            rows.append(["TOTAL RECEITAS",brl(dre_total_receitas),"100.0%"])
            story.append(_tbl(rows,[320,150,130],[("FONTNAME",(0,len(rows)-1),(-1,len(rows)-1),"Helvetica-Bold"),("BACKGROUND",(0,len(rows)-1),(-1,len(rows)-1),colors.HexColor("#e8f5e9"))]))
        else: story.append(Paragraph("Nenhuma receita realizada. Marque os lancamentos como Recebido para aparecerem na DRE.",normal))
        story.append(Spacer(1,10))
        story.append(Paragraph("DESPESAS - itens marcados como Pago", bold9)); story.append(Spacer(1,4))
        if dre_despesas:
            rows=[["Categoria","Valor","% Receitas"]]
            for cat,val in dre_despesas: rows.append([cat,brl(val),f"{val/dre_total_receitas*100:.1f}%" if dre_total_receitas else "0%"])
            rows.append(["TOTAL DESPESAS",brl(dre_total_despesas),f"{dre_total_despesas/dre_total_receitas*100:.1f}%" if dre_total_receitas else "0%"])
            story.append(_tbl(rows,[320,150,130],[("FONTNAME",(0,len(rows)-1),(-1,len(rows)-1),"Helvetica-Bold"),("BACKGROUND",(0,len(rows)-1),(-1,len(rows)-1),colors.HexColor("#fdecea"))]))
        else: story.append(Paragraph("Nenhuma despesa paga. Marque os lancamentos como Pago para aparecerem na DRE.",normal))
        story.append(Spacer(1,10))
        res_row = [["RESULTADO LIQUIDO", brl(dre_resultado), f"Margem: {dre_margem:.1f}%"]]
        story.append(_tbl([["","Valor",""]]+res_row,[240,150,210],[("FONTNAME",(0,1),(-1,1),"Helvetica-Bold"),("FONTSIZE",(0,1),(-1,1),10),("BACKGROUND",(0,1),(-1,1),colors.HexColor("#27AE60") if dre_resultado>=0 else colors.HexColor("#C0392B")),("TEXTCOLOR",(0,1),(-1,1),colors.white)]))

    elif tipo == "fluxo":
        tr, td, res = totais(dados_financeiro)
        saldo_final = fluxo[-1]["saldo"] if fluxo else 0.0
        story.append(_kpi_table([
            ("Total Entradas", brl(tr),          "#27AE60"),
            ("Total Saidas",   brl(td),          "#C0392B"),
            ("Resultado",      brl(res),         "#27AE60" if res>=0 else "#C0392B"),
            ("Saldo Final",    brl(saldo_final), "#27AE60" if saldo_final>=0 else "#C0392B"),
        ]))
        story.append(Spacer(1,10))
        if fluxo:
            rows=[["Data","Tipo","Descricao","Categoria","Entrada (+)","Saida (-)","Saldo Acum.","Status"]]
            for d in fluxo:
                s=d.get("saldo",0)
                rows.append([d.get("vencimento") or "",d.get("tipo"),Paragraph(str(d.get("descricao") or ""),small),d.get("categoria") or "",brl(d["entrada"]) if d.get("entrada") else "-",brl(d["saida"]) if d.get("saida") else "-",brl(s),d.get("status") or ""])
            story.append(_tbl(rows,[72,55,188,88,78,78,88,58],[("TEXTCOLOR",(6,1),(6,-1),colors.HexColor("#1a6e39"))]))
        else: story.append(Paragraph("Nenhum lancamento encontrado no periodo.",normal))

    else:
        tr, td, res = totais(dados_financeiro)
        n_aberto = sum(1 for d in dados_financeiro if (d.get("status") or "")=="aberto")
        story.append(_kpi_table([
            ("Total Receitas", brl(tr),             "#27AE60"),
            ("Total Despesas", brl(td),             "#C0392B"),
            ("Resultado",      brl(res),            "#27AE60" if res>=0 else "#C0392B"),
            ("Em Aberto",      f"{n_aberto} item(s)","#F39C12"),
        ]))
        story.append(Spacer(1,10))
        if dados_financeiro:
            rows=[["Tipo","Descricao","Categoria","Valor","Vencimento","Status","Centro Custo","Forma Pgto"]]
            for d in dados_financeiro:
                rows.append([d.get("tipo"),Paragraph(str(d.get("descricao") or ""),small),d.get("categoria") or "",brl(d.get("valor") or 0),d.get("vencimento") or "",d.get("status") or "",d.get("centro_custo") or "",d.get("forma_pagamento") or ""])
            story.append(_tbl(rows,[52,198,88,82,72,62,90,68]))
        else: story.append(Paragraph("Nenhum lancamento encontrado no periodo.",normal))

    doc.build(story)
    bio.seek(0)
    disp = "inline" if preview else "attachment"
    return StreamingResponse(bio, media_type="application/pdf", headers={"Content-Disposition": f"{disp}; filename={filename('pdf')}"})



@app.get("/api/migrate2")
def migrate2(admin=Depends(requer_admin)):
    return {"ok": True, "msg": "Migração manual desativada por segurança. As tabelas são verificadas na inicialização."}


@app.post("/api/chat")
def chat(payload: dict, usuario=Depends(get_usuario)):
    return {"resposta": "Agente IA em revisão técnica. Os dados financeiros seguem disponíveis nos relatórios, DRE e dashboard."}


if os.path.isdir("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
def root():
    index_path = os.path.join("static", "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return {"ok": True, "app": "Financeiro Academia"}

# ============ NOTIFICAÇÕES =====@app.get("/api/notificacoes")
def get_notificacoes(usuario=Depends(get_usuario)):
    from datetime import timedelta
    conn = get_db()
    hoje = datetime.today().date()
    em3 = (hoje + timedelta(days=3)).isoformat()
    hoje_str = hoje.isoformat()
    fr = " AND restrita=0" if not ve_restritas(usuario) else ""

    vencendo = fetchall_dict(conn,
        f"SELECT *,'pagar' as tipo FROM contas_pagar WHERE status='aberto' AND vencimento <= ? AND vencimento >= ?{fr} ORDER BY vencimento",
        (em3, hoje_str))
    vencendo += fetchall_dict(conn,
        f"SELECT *,'receber' as tipo FROM contas_receber WHERE status='aberto' AND vencimento <= ? AND vencimento >= ?{fr} ORDER BY vencimento",
        (em3, hoje_str))

    vencidas = fetchall_dict(conn,
        f"SELECT *,'pagar' as tipo FROM contas_pagar WHERE status='aberto' AND vencimento < ?{fr} ORDER BY vencimento",
        (hoje_str,))
    vencidas += fetchall_dict(conn,
        f"SELECT *,'receber' as tipo FROM contas_receber WHERE status='aberto' AND vencimento < ?{fr} ORDER BY vencimento",
        (hoje_str,))

    conn.close()
    total = len(vencendo) + len(vencidas)
    return {
        "total": total,
        "vencendo": [dict(r) for r in vencendo],
        "vencidas": [dict(r) for r in vencidas]
    }


# ============ NOTIFICACOES POR EMAIL + WHATSAPP =====import smtplib as _smtplib
from email.mime.multipart import MIMEMultipart as _MIMEMultipart
from email.mime.text import MIMEText as _MIMEText
import urllib.request as _urllib_req
import json as _json

def _fmt_brl(v):
    try:
        return "R$ {:,.2f}".format(float(v)).replace(",","X").replace(".",",").replace("X",".")
    except:
        return str(v)

def _get_vencimentos():
    conn = get_db()
    hoje = datetime.today().date()
    em3 = (hoje + timedelta(days=3)).isoformat()
    hoje_str = hoje.isoformat()
    vencendo = fetchall_dict(conn,
        "SELECT *,'pagar' as tipo FROM contas_pagar WHERE status='aberto' AND vencimento <= ? AND vencimento >= ? ORDER BY vencimento",
        (em3, hoje_str))
    vencendo += fetchall_dict(conn,
        "SELECT *,'receber' as tipo FROM contas_receber WHERE status='aberto' AND vencimento <= ? AND vencimento >= ? ORDER BY vencimento",
        (em3, hoje_str))
    vencidas = fetchall_dict(conn,
        "SELECT *,'pagar' as tipo FROM contas_pagar WHERE status='aberto' AND vencimento < ? ORDER BY vencimento",
        (hoje_str,))
    vencidas += fetchall_dict(conn,
        "SELECT *,'receber' as tipo FROM contas_receber WHERE status='aberto' AND vencimento < ? ORDER BY vencimento",
        (hoje_str,))
    conn.close()
    return hoje, vencendo, vencidas

def enviar_email_notificacoes():
    email_from = os.getenv("EMAIL_FROM")
    email_pass = os.getenv("EMAIL_PASSWORD")
    if not email_from or not email_pass:
        return
    try:
        conn = get_db()
        hoje, vencendo, vencidas = _get_vencimentos()
        total = len(vencendo) + len(vencidas)
        if total == 0:
            return
        fixos = [e.strip() for e in os.getenv("EMAIL_TO","").split(",") if e.strip()]
        usuarios_notif = fetchall_dict(conn,
            "SELECT email FROM usuarios WHERE perfil IN ('admin','gestor','gerente') AND ativo=1")
        conn.close()
        destinatarios = list(set(fixos + [u["email"] for u in usuarios_notif if u.get("email")]))
        if not destinatarios:
            return
        def _row(c, cor):
            tipo_txt = "A Pagar" if c.get("tipo")=="pagar" else "A Receber"
            return (f"<tr><td style='padding:6px 10px;border-bottom:1px solid #eee'>{c.get('descricao','')}</td>"
                    f"<td style='padding:6px 10px;border-bottom:1px solid #eee;color:#666'>{c.get('vencimento','')}</td>"
                    f"<td style='padding:6px 10px;border-bottom:1px solid #eee;color:{cor};font-weight:bold'>{_fmt_brl(c.get('valor',0))}</td>"
                    f"<td style='padding:6px 10px;border-bottom:1px solid #eee;color:#999;font-size:11px'>{tipo_txt}</td></tr>")
        rows_v = "".join(_row(c,"#E74C3C") for c in vencidas)
        rows_p = "".join(_row(c,"#F59E0B") for c in vencendo)
        sec_v = (f"<h3 style='color:#E74C3C'>&#9888; Contas Vencidas ({len(vencidas)})</h3>"
                 f"<table width='100%' style='border-collapse:collapse;font-size:13px'>{rows_v}</table>") if vencidas else ""
        sec_p = (f"<h3 style='color:#F59E0B'>&#128276; Vencem nos Proximos 3 Dias ({len(vencendo)})</h3>"
                 f"<table width='100%' style='border-collapse:collapse;font-size:13px'>{rows_p}</table>") if vencendo else ""
        data_str = hoje.strftime("%d/%m/%Y")
        html_body = (f"<div style='font-family:Arial,sans-serif;max-width:640px;margin:0 auto'>"
                     f"<div style='background:#C0392B;padding:20px;color:#fff'><h2 style='margin:0'>Body Finance - Alertas</h2><p style='margin:4px 0 0'>{data_str}</p></div>"
                     f"<div style='padding:24px'>{sec_v}{sec_p}"
                     f"<p style='margin:20px 0 0;font-size:12px;color:#999'><a href='https://bodyfinance.up.railway.app/' style='color:#C0392B'>Acessar sistema</a></p>"
                     f"</div></div>")
        msg = _MIMEMultipart("alternative")
        msg["Subject"] = f"[Body Finance] {total} alerta(s) - {data_str}"
        msg["From"] = f"Body Finance <{email_from}>"
        msg["To"] = ", ".join(destinatarios)
        msg.attach(_MIMEText(html_body, "html", "utf-8"))
        with _smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15) as srv:
            srv.login(email_from, email_pass)
            srv.sendmail(email_from, destinatarios, msg.as_string())
        print(f"[EMAIL] Enviado para {len(destinatarios)} destinatario(s)")
    except Exception as e:
        print(f"[EMAIL] Erro: {e}")

def enviar_whatsapp_notificacoes():
    instance_id = os.getenv("ZAPI_INSTANCE_ID")
    token = os.getenv("ZAPI_INSTANCE_TOKEN")
    client_token = os.getenv("ZAPI_CLIENT_TOKEN")
    phones_str = os.getenv("WHATSAPP_PHONES","")
    if not instance_id or not token or not phones_str:
        return
    try:
        hoje, vencendo, vencidas = _get_vencimentos()
        total = len(vencendo) + len(vencidas)
        if total == 0:
            return
        data_str = hoje.strftime("%d/%m/%Y")
        linhas = [f"*Body Finance - Alertas* ({data_str})"]
        if vencidas:
            linhas.append(f"\n*\u26a0\ufe0f Contas Vencidas ({len(vencidas)}):*")
            for c in vencidas:
                linhas.append(f"  - {c.get('descricao','')} | {c.get('vencimento','')} | {_fmt_brl(c.get('valor',0))}")
        if vencendo:
            linhas.append(f"\n*\U0001f514 Vencem em ate 3 dias ({len(vencendo)}):*")
            for c in vencendo:
                linhas.append(f"  - {c.get('descricao','')} | {c.get('vencimento','')} | {_fmt_brl(c.get('valor',0))}")
        linhas.append("\nhttps://bodyfinance.up.railway.app/")
        mensagem = "\n".join(linhas)
        phones = [p.strip() for p in phones_str.split(",") if p.strip()]
        url = f"https://api.z-api.io/instances/{instance_id}/send-text"
        for phone in phones:
            payload = _json.dumps({"phone": phone, "message": mensagem}).encode("utf-8")
            req = _urllib_req.Request(url, data=payload, method="POST")
            req.add_header("Content-Type", "application/json")
            req.add_header("Authorization", f"Bearer {token}")
            req.add_header("Client-Token", client_token)
            with _urllib_req.urlopen(req, timeout=15) as resp:
                print(f"[WHATSAPP] Enviado para {phone}: {resp.status}")
    except Exception as e:
        print(f"[WHATSAPP] Erro: {e}")

def _enviar_todas_notificacoes():
    enviar_email_notificacoes()
    enviar_whatsapp_notificacoes()

_notif_scheduler = BackgroundScheduler(timezone="America/Sao_Paulo")
_notif_scheduler.add_job(_enviar_todas_notificacoes, "cron", hour=8, minute=0)

@app.on_event("startup")
def start_notif_scheduler():
    _notif_scheduler.start()
    print("[NOTIF] Scheduler iniciado - envio diario as 08:00 (email + whatsapp)")

@app.on_event("shutdown")
def stop_notif_scheduler():
    if _notif_scheduler.running:
        _notif_scheduler.shutdown(wait=False)

@app.get("/{full_path:path}")
def catch_all(full_path: str)