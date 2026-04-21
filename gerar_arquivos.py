import os

relatorios = '''import sqlite3,csv,io,datetime

def get_db():
    conn=sqlite3.connect("financeiro.db")
    conn.row_factory=sqlite3.Row
    return conn

def fmtD(d):
    try:return datetime.date.fromisoformat(d).strftime("%d/%m/%Y")
    except:return d

def fmtR_str(v):
    try:
        return "R$ "+"{:,.2f}".format(float(v)).replace(",","X").replace(".",",").replace("X",".")
    except:return str(v)

def periodo_datas(periodo,data_ini=None,data_fim=None):
    hoje=datetime.date.today()
    if periodo=="personalizado" and data_ini and data_fim:return data_ini,data_fim
    if periodo=="diario":return hoje.isoformat(),hoje.isoformat()
    if periodo=="semanal":
        ini=hoje-datetime.timedelta(days=hoje.weekday())
        return ini.isoformat(),(ini+datetime.timedelta(days=6)).isoformat()
    ini=hoje.replace(day=1)
    if hoje.month==12:fim=hoje.replace(year=hoje.year+1,month=1,day=1)-datetime.timedelta(days=1)
    else:fim=hoje.replace(month=hoje.month+1,day=1)-datetime.timedelta(days=1)
    return ini.isoformat(),fim.isoformat()

def buscar_pagar(ini,fim,ve_restrita=True,status=None):
    conn=get_db()
    conds=["vencimento >= ?","vencimento <= ?"]
    params=[ini,fim]
    if not ve_restrita:conds.append("restrita=0")
    rows=conn.execute("SELECT * FROM contas_pagar WHERE "+" AND ".join(conds)+" ORDER BY vencimento",params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def buscar_receber(ini,fim,ve_restrita=True,status=None):
    conn=get_db()
    conds=["vencimento >= ?","vencimento <= ?"]
    params=[ini,fim]
    if not ve_restrita:conds.append("restrita=0")
    rows=conn.execute("SELECT * FROM contas_receber WHERE "+" AND ".join(conds)+" ORDER BY vencimento",params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def buscar_inadimplencia(ve_restrita=True):
    conn=get_db()
    hoje=datetime.date.today().isoformat()
    fr="" if ve_restrita else " AND restrita=0"
    rows=conn.execute("SELECT * FROM contas_receber WHERE status='aberto' AND vencimento < ?"+fr+" ORDER BY vencimento",(hoje,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def gerar_csv(headers,rows,totals=None):
    out=io.StringIO()
    w=csv.writer(out)
    w.writerow(headers)
    for r in rows:w.writerow(r)
    if totals:w.writerow([]);w.writerow(totals)
    return out.getvalue().encode("utf-8-sig")

def gerar_csv_pagar(rows):
    return gerar_csv(["Desc","Cat","Valor","Venc","Status"],
        [(r["desc"],r["categoria"],"{:.2f}".format(r["valor"]),fmtD(r["vencimento"]),r["status"]) for r in rows],
        ["TOTAL","","{:.2f}".format(sum(r["valor"] for r in rows)),"",""])

def gerar_csv_receber(rows):return gerar_csv_pagar(rows)

def gerar_csv_fluxo(pagar,receber):
    rows=[("E",r["desc"],r["categoria"],"{:.2f}".format(r["valor"]),fmtD(r["vencimento"]),r["status"]) for r in receber]
    rows+=[("S",r["desc"],r["categoria"],"{:.2f}".format(r["valor"]),fmtD(r["vencimento"]),r["status"]) for r in pagar]
    tot=sum(r["valor"] for r in receber)-sum(r["valor"] for r in pagar)
    return gerar_csv(["Tipo","Desc","Cat","Valor","Venc","Status"],rows,["RES","","","{:.2f}".format(tot),"",""])

def gerar_csv_dre(pagar,receber):
    receita=sum(r["valor"] for r in receber)
    cats={}
    for r in pagar:cats[r["categoria"]]=cats.get(r["categoria"],0)+r["valor"]
    td=sum(cats.values())
    rows=[("Receita","{:.2f}".format(receita))]
    for cat,val in sorted(cats.items(),key=lambda x:-x[1]):rows.append((cat,"-{:.2f}".format(val)))
    rows+=[("Total","-{:.2f}".format(td)),("Resultado","{:.2f}".format(receita-td))]
    return gerar_csv(["Item","Valor"],rows)

def gerar_csv_inadimplencia(rows):
    hoje=datetime.date.today()
    return gerar_csv(["Desc","Cat","Valor","Venc","Dias"],
        [(r["desc"],r["categoria"],"{:.2f}".format(r["valor"]),fmtD(r["vencimento"]),(hoje-datetime.date.fromisoformat(r["vencimento"])).days) for r in rows],
        ["TOTAL","","{:.2f}".format(sum(r["valor"] for r in rows)),"",""])

def gerar_csv_resumo(pagar,receber,ini,fim):
    tr=sum(r["valor"] for r in receber)
    tp=sum(r["valor"] for r in pagar)
    rr=sum(r["valor"] for r in receber if r["status"]=="recebido")
    pr=sum(r["valor"] for r in pagar if r["status"]=="pago")
    return gerar_csv(["Ind","Prev","Real","Dif"],[
        ("Rec","{:.2f}".format(tr),"{:.2f}".format(rr),"{:.2f}".format(rr-tr)),
        ("Desp","{:.2f}".format(tp),"{:.2f}".format(pr),"{:.2f}".format(pr-tp)),
        ("Res","{:.2f}".format(tr-tp),"{:.2f}".format(rr-pr),"{:.2f}".format((rr-pr)-(tr-tp)))])

def gerar_xlsx(sheets):
    import openpyxl,io as _io
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    wb=openpyxl.Workbook()
    wb.remove(wb.active)
    thin=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    for sname,headers,rows,totals in sheets:
        ws=wb.create_sheet(title=sname[:31])
        for col,h in enumerate(headers,1):
            c=ws.cell(row=1,column=col,value=h)
            c.font=Font(name="Arial",bold=True,color="FFFFFFFF",size=10)
            c.fill=PatternFill("solid",fgColor="FF1A1A1A")
            c.alignment=Alignment(horizontal="center")
            c.border=thin
        for ri,row in enumerate(rows,2):
            fill=PatternFill("solid",fgColor="FFFFFFFF") if ri%2==0 else PatternFill("solid",fgColor="FFFAFAF8")
            for col,val in enumerate(row,1):
                c=ws.cell(row=ri,column=col,value=val)
                c.font=Font(name="Arial",size=10)
                c.border=thin
                c.fill=fill
                if isinstance(val,float):c.number_format="#,##0.00";c.alignment=Alignment(horizontal="right")
        if totals:
            tr2=len(rows)+2
            for col,val in enumerate(totals,1):
                c=ws.cell(row=tr2,column=col,value=val)
                c.font=Font(name="Arial",bold=True,color="FFFFFFFF",size=10)
                c.fill=PatternFill("solid",fgColor="FF0F6E56")
                c.border=thin
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(len(str(c.value or ""))+4 for c in col),40)
    buf=_io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def build_xlsx_pagar(rows):
    return gerar_xlsx([("Pagar",["Desc","Cat","Valor","Venc","Status"],
        [(r["desc"],r["categoria"],float(r["valor"]),fmtD(r["vencimento"]),r["status"]) for r in rows],
        ["TOTAL","",float(sum(r["valor"] for r in rows)),"",""])])

def build_xlsx_receber(rows):
    return gerar_xlsx([("Receber",["Desc","Cat","Valor","Venc","Status"],
        [(r["desc"],r["categoria"],float(r["valor"]),fmtD(r["vencimento"]),r["status"]) for r in rows],
        ["TOTAL","",float(sum(r["valor"] for r in rows)),"",""])])

def build_xlsx_fluxo(pagar,receber):
    data=[("E",r["desc"],r["categoria"],float(r["valor"]),fmtD(r["vencimento"]),r["status"]) for r in receber]
    data+=[("S",r["desc"],r["categoria"],float(r["valor"]),fmtD(r["vencimento"]),r["status"]) for r in pagar]
    tot=float(sum(r["valor"] for r in receber)-sum(r["valor"] for r in pagar))
    return gerar_xlsx([("Fluxo",["Tipo","Desc","Cat","Valor","Venc","Status"],data,["RES","","",tot,"",""])])

def build_xlsx_dre(pagar,receber):
    receita=sum(r["valor"] for r in receber)
    cats={}
    for r in pagar:cats[r["categoria"]]=cats.get(r["categoria"],0)+r["valor"]
    td=sum(cats.values())
    res=receita-td
    m=round(res/receita*100,1) if receita else 0
    data=[("Receita",float(receita),"")]
    for c,v in sorted(cats.items(),key=lambda x:-x[1]):data.append((c,float(-v),""))
    data+=[("Total",float(-td),""),("Resultado",float(res),""),("Margem",str(m)+"%","")]
    return gerar_xlsx([("DRE",["Item","Valor","Obs"],data,None)])

def build_xlsx_inadimplencia(rows):
    hoje=datetime.date.today()
    data=[(r["desc"],r["categoria"],float(r["valor"]),fmtD(r["vencimento"]),(hoje-datetime.date.fromisoformat(r["vencimento"])).days) for r in rows]
    return gerar_xlsx([("Inadimpl",["Desc","Cat","Valor","Venc","Dias"],data,["TOTAL","",float(sum(r["valor"] for r in rows)),"",""])])

def build_xlsx_resumo(pagar,receber,ini,fim):
    tr=sum(r["valor"] for r in receber)
    tp=sum(r["valor"] for r in pagar)
    rr=sum(r["valor"] for r in receber if r["status"]=="recebido")
    pr=sum(r["valor"] for r in pagar if r["status"]=="pago")
    return gerar_xlsx([("Resumo",["Ind","Prev","Real","Dif"],[
        ("Rec",float(tr),float(rr),float(rr-tr)),
        ("Desp",float(tp),float(pr),float(pr-tp)),
        ("Res",float(tr-tp),float(rr-pr),float((rr-pr)-(tr-tp)))],None)])

def gerar_pdf_html(titulo,periodo_str,tabelas):
    rows_html=""
    for sub,headers,rows,totals in tabelas:
        rows_html+="<h2>"+sub+"</h2><table><thead><tr>"
        rows_html+="".join("<th>"+h+"</th>" for h in headers)+"</tr></thead><tbody>"
        for row in rows:rows_html+="<tr>"+"".join("<td>"+str(v)+"</td>" for v in row)+"</tr>"
        if totals:rows_html+='<tr class="total">'+"".join("<td>"+str(v)+"</td>" for v in totals)+"</tr>"
        rows_html+="</tbody></table>"
    now=datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    return "<!DOCTYPE html><html><head><meta charset=UTF-8><title>"+titulo+"</title><style>body{font-family:Arial;font-size:11px;margin:20px}h1{color:#0F6E56;border-bottom:2px solid #0F6E56}table{width:100%;border-collapse:collapse;margin-bottom:16px}th{background:#1a1a1a;color:#fff;padding:6px 8px;text-align:left}td{padding:6px 8px;border-bottom:1px solid #e5e5e0}tr:nth-child(even) td{background:#fafaf8}tr.total td{background:#0F6E56;color:#fff;font-weight:bold}@media print{body{margin:10mm}}</style></head><body><h1>"+titulo+"</h1><p>Periodo: "+periodo_str+" | Gerado: "+now+"</p>"+rows_html+"</body></html>"
'''

open("relatorios.py","w",encoding="utf-8").write(relatorios)
print("relatorios.py criado!")